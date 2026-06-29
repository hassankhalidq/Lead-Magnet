"""
EDGAR Form D Lead Finder
-------------------------
Finds startups that recently filed SEC Form D (private capital raise
disclosure), filtered by state and industry keyword.

This is a free, public-data tool. It does NOT scrape personal emails or
phone numbers. It surfaces the company name, filing date, and the named
executives/related persons who SIGNED the public filing (legally required
disclosure), plus a direct link to the filing itself on sec.gov. You take
it from there: look up that named person on LinkedIn, the company site, or
wherever they've made themselves reachable.

Setup (one time):
    pip install edgartools

Usage:
    python edgar_lead_finder.py --state Texas --keyword energy --start-date 2026-01-01 --end-date 2026-06-01
    python edgar_lead_finder.py --state California "New York" Massachusetts Texas Colorado Washington --keyword energy "climate tech" renewable --start-date 2026-01-01 --end-date 2026-06-01 --limit 100

Notes:
- SEC EDGAR is a free, public US government database. No API key required.
- The library identifies itself to SEC's servers per their usage policy —
  edit IDENTITY_EMAIL below to your real contact info before running, since
  SEC asks for this (it's not optional, it's how they let you in politely).
- "state" filters by the issuer's principal office state, not the founder's
  personal location.
- "keyword" is matched against the company name and industry description.
  Form D's industry field is broad (e.g. "Other Energy"), so a keyword like
  "energy" or "climate" is more reliable than a narrow SIC code.

Keyword guidance (learned from testing):
- GOOD single words: energy, solar, hydrogen, battery, carbon — specific
  enough to mostly avoid noise, common enough to actually appear in filings.
- BAD single words: "storage" alone pulls in real-estate self-storage
  businesses (literal storage units), not energy storage startups. Use a
  two-word phrase instead if you want that niche, e.g. "energy storage" or
  "battery storage".
- BAD phrases: multi-word marketing buzzwords like "climate tech" almost
  never appear verbatim inside a legal filing's text, so they return 0
  results across most states. Filings use plain technical/business language,
  not pitch-deck language — search for what the company actually does
  (solar, hydrogen, carbon capture), not the category investors use to
  describe it.
- When trying a new keyword, skim a few of its results before trusting the
  count — a high hit count can mean a real niche OR an unrelated industry
  that happens to share the word.
"""

import argparse
import csv
import os
from datetime import datetime
from edgar import search_filings, set_identity
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

IDENTITY_EMAIL = "Lead Research your_email_here@example.com"  # <-- replace with your real name/email LOCALLY only. Do not commit your real email to a public repo.

SEEN_LEADS_FILE = "seen_leads.xlsx"  # lives next to this script; one tab per state, tracks every lead ever shown


def estimate_stage(total_offering_amount):
    """
    Rough stage guess based on the dollar amount being raised, per Form D's
    own disclosed total_offering_amount field. This is a heuristic, not a
    certainty — a $2M raise could be a small seed round or a bridge, always
    verify with a real search before treating this as fact.
    """
    if not total_offering_amount:
        return "Unknown"
    try:
        amount = float(str(total_offering_amount).replace(",", "").replace("$", ""))
    except (ValueError, TypeError):
        return "Unknown"

    if amount <= 1_000_000:
        return "Likely pre-seed"
    elif amount <= 5_000_000:
        return "Likely seed"
    elif amount <= 20_000_000:
        return "Likely Series A"
    else:
        return "Likely Series B+ (probably too late-stage for Ignite/Boost)"


def days_since(filing_date) -> int:
    """Returns how many days ago a filing was made, for recency sorting/flagging."""
    try:
        parsed = filing_date if isinstance(filing_date, datetime) else datetime.strptime(str(filing_date), "%Y-%m-%d")
        return (datetime.now() - parsed).days
    except (ValueError, TypeError):
        return 9999  # unknown/unparseable date — sort to the bottom rather than crash


def recency_flag(days: int) -> str:
    if days <= 30:
        return "Fresh"
    elif days <= 90:
        return "Recent"
    elif days <= 180:
        return "Aging"
    else:
        return "Old"


COLUMNS = ["company", "first_seen_date", "filing_date", "days_since_filing", "recency", "matched_keyword", "estimated_stage", "filing_link", "status"]


def _sanitize_sheet_name(state: str) -> str:
    """Excel sheet names can't exceed 31 chars or contain certain symbols."""
    name = state.strip()[:31]
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, '')
    return name or "Unknown"


def load_seen_leads():
    """Returns {state_sheet_name: set of lowercased company names} already shown in past runs."""
    if not os.path.exists(SEEN_LEADS_FILE):
        return {}
    wb = load_workbook(SEEN_LEADS_FILE)
    seen = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        companies = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                companies.add(str(row[0]).strip().lower())
        seen[sheet_name] = companies
    return seen


def append_to_seen_leads(results):
    """Appends newly-shown leads to seen_leads.xlsx, one sheet per state.
    Creates the workbook/sheets if they don't exist yet, otherwise appends
    to existing sheets without disturbing rows you've already edited."""
    if os.path.exists(SEEN_LEADS_FILE):
        wb = load_workbook(SEEN_LEADS_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default blank sheet; we add real ones below

    by_state = {}
    for r in results:
        by_state.setdefault(r["matched_state"], []).append(r)

    for state, rows in by_state.items():
        sheet_name = _sanitize_sheet_name(state)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            sheet.append(COLUMNS)
            for cell in sheet[1]:
                cell.font = Font(bold=True)

        for r in rows:
            sheet.append(
                [
                    r["company"],
                    datetime.now().strftime("%Y-%m-%d"),
                    str(r["filing_date"]),
                    r["days_since_filing"],
                    r["recency"],
                    r["matched_keyword"],
                    r["estimated_stage"],
                    r["filing_link"],
                    "Not yet",  # you edit this column yourself: Contacted / Pass / etc.
                ]
            )

        for col_idx, width in enumerate([35, 14, 12, 16, 10, 16, 38, 60, 12], start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = width

    wb.save(SEEN_LEADS_FILE)


def _search_one_state(state: str, keyword: str, start_date: str, end_date: str, limit: int):
    """Runs one server-side search for a single state and returns parsed,
    fund-filtered results (this is the original single-state logic)."""
    query = f"{state} {keyword}"
    print(f"  Searching '{query}'...")

    search_results = search_filings(
        query,
        forms="D",
        start_date=start_date,
        end_date=end_date,
        limit=limit,
    )
    print(f"    -> {len(search_results)} raw results")

    results = []
    excluded_fund_count = 0
    fund_indicators = [
        " lp", " l.p.", "fund", "partners", "capital partners",
        "private capital", "investment fund",
    ]

    for r in search_results:
        try:
            filing = r.get_filing()
            data = filing.obj()

            company_lower = (filing.company or "").lower()
            # Skip investment funds / PE/LP vehicles — these raise capital to
            # invest in OTHER companies, they are not operating startups
            # looking for product funding (the Cephyron-relevant kind).
            if any(indicator in company_lower for indicator in fund_indicators):
                excluded_fund_count += 1
                continue

            related_persons = getattr(data, "related_persons", []) or []
            names = [
                f"{getattr(p, 'first_name', '') or ''} {getattr(p, 'last_name', '') or ''}".strip()
                for p in related_persons
            ]
            names = [n for n in names if n]  # drop any empty strings

            offering_data = getattr(data, "offering_data", None)
            sales_amounts = getattr(offering_data, "offering_sales_amounts", None) if offering_data else None
            total_offering_amount = getattr(sales_amounts, "total_offering_amount", None) if sales_amounts else None

            days_old = days_since(filing.filing_date)

            results.append(
                {
                    "company": filing.company,
                    "filing_date": filing.filing_date,
                    "named_executives": names or ["(not parsed — open filing link)"],
                    "filing_link": filing.filing_url,
                    "matched_state": state,
                    "matched_keyword": keyword,
                    "total_offering_amount": total_offering_amount,
                    "estimated_stage": estimate_stage(total_offering_amount),
                    "days_since_filing": days_old,
                    "recency": recency_flag(days_old),
                }
            )
        except Exception:
            # Some filings are malformed or use older schemas — skip, don't crash.
            continue

    if excluded_fund_count:
        print(f"    (filtered out {excluded_fund_count} fund/LP-style filings)")

    return results


def find_form_d_leads(states, keywords, start_date: str, end_date: str, limit: int = 25):
    """
    Runs one server-side search per (state, keyword) combination, since
    EDGAR's search doesn't support multi-value queries directly. Then
    combines, de-duplicates by company name, and sorts by filing recency
    (freshest first) so the most time-sensitive leads surface at the top.
    """
    set_identity(IDENTITY_EMAIL)

    print(f"Searching Form D filings across {len(states)} state(s) x {len(keywords)} keyword(s) between {start_date} and {end_date}...\n")

    all_results = []
    for state in states:
        for keyword in keywords:
            state_results = _search_one_state(state, keyword, start_date, end_date, limit)
            all_results.extend(state_results)

    # De-duplicate by company name (same company can surface across more
    # than one state/keyword combination)
    seen_companies = set()
    deduped = []
    for r in all_results:
        key = r["company"].strip().lower() if r["company"] else ""
        if key and key not in seen_companies:
            seen_companies.add(key)
            deduped.append(r)

    duplicates_removed = len(all_results) - len(deduped)
    if duplicates_removed:
        print(f"\n(Removed {duplicates_removed} duplicate company entries across states/keywords.)")

    # Skip anything already shown in a previous run of this script (checked per-state sheet)
    already_seen = load_seen_leads()
    new_results = []
    skipped_count = 0
    for r in deduped:
        sheet_name = _sanitize_sheet_name(r["matched_state"])
        seen_in_this_state = already_seen.get(sheet_name, set())
        if r["company"].strip().lower() in seen_in_this_state:
            skipped_count += 1
        else:
            new_results.append(r)

    if skipped_count:
        print(f"(Skipped {skipped_count} companies already seen in a previous run — check {SEEN_LEADS_FILE} for their status.)")

    # Freshest filings first — the most time-sensitive leads surface at the top
    new_results.sort(key=lambda r: r["days_since_filing"])

    if new_results:
        append_to_seen_leads(new_results)

    return new_results


def print_results(results):
    if not results:
        print("\nNo new matching filings found (or everything found was already seen in a previous run — check seen_leads.csv).")
        return

    print(f"\nFound {len(results)} NEW matching filings (sorted freshest first):\n")
    for i, r in enumerate(results, 1):
        amount_str = f"${r['total_offering_amount']}" if r['total_offering_amount'] else "amount not disclosed"
        print(f"{i}. {r['company']}  [{r['matched_state']}]  ({r['recency']}, {r['days_since_filing']}d ago)")
        print(f"   Filed: {r['filing_date']}  |  Raising: {amount_str}  |  Stage guess: {r['estimated_stage']}  |  Matched on: '{r['matched_keyword']}'")
        print(f"   Named on filing: {', '.join(r['named_executives'])}")
        print(f"   Filing link: {r['filing_link']}")
        print()

    print(f"All {len(results)} leads above have been added to {SEEN_LEADS_FILE}.")
    print(f"Open that file to track status (Contacted / Pass / etc.) — future runs will skip anything already in it.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Find recent SEC Form D filings across one or more states and keywords (fast, server-side search).")
    parser.add_argument("--state", required=True, nargs="+", help="One or more state names, e.g. --state Texas California \"New York\"")
    parser.add_argument("--keyword", required=True, nargs="+", help="One or more keywords, e.g. --keyword energy \"climate tech\" renewable")
    parser.add_argument("--start-date", required=True, help="Start date, format YYYY-MM-DD")
    parser.add_argument("--end-date", required=True, help="End date, format YYYY-MM-DD")
    parser.add_argument("--limit", type=int, default=25, help="Max results PER STATE PER KEYWORD (SEC caps this around 100)")
    args = parser.parse_args()

    results = find_form_d_leads(
        states=args.state,
        keywords=args.keyword,
        start_date=args.start_date,
        end_date=args.end_date,
        limit=args.limit,
    )
    print_results(results)
